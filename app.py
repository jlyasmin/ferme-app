import json, os, hashlib, uuid, re
from datetime import date
from http.server import BaseHTTPRequestHandler, HTTPServer
from urllib.parse import urlparse

# Railway garde les fichiers dans /app — jamais effacés
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
USERS_FILE = os.path.join(BASE_DIR, "users.json")
DATA_FILE  = os.path.join(BASE_DIR, "data.json")
sessions   = {}

def load(f):
    try:
        if os.path.exists(f):
            with open(f, "r", encoding="utf-8") as fp:
                return json.load(fp)
    except: pass
    return {}

def save(f, v):
    with open(f, "w", encoding="utf-8") as fp:
        json.dump(v, fp, ensure_ascii=False, indent=2)

def sha(s): return hashlib.sha256(s.encode()).hexdigest()
def new_session(u):
    t = str(uuid.uuid4()); sessions[t] = u; return t
def get_user(token): return sessions.get(token)

def init():
    if not os.path.exists(USERS_FILE): save(USERS_FILE, {})
    if not os.path.exists(DATA_FILE):  save(DATA_FILE, {})

CSS = """
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
     background:linear-gradient(135deg,#1B5E20,#4CAF50);
     min-height:100vh;display:flex;align-items:center;justify-content:center;}
.card{background:white;border-radius:18px;padding:40px 36px;width:380px;
      box-shadow:0 24px 64px rgba(0,0,0,.22);}
.logo{text-align:center;margin-bottom:28px;}
.em{font-size:52px;}
h1{font-size:22px;font-weight:700;color:#1B5E20;margin-top:8px;}
.sub{font-size:13px;color:#999;margin-top:4px;}
.tabs{display:flex;border:1.5px solid #e0e0e0;border-radius:10px;
      overflow:hidden;margin-bottom:22px;}
.tab{flex:1;padding:10px;text-align:center;font-size:14px;font-weight:600;
     cursor:pointer;background:white;color:#888;border:none;transition:all .2s;}
.tab.active{background:#2E7D32;color:white;}
.panel{display:none;}.panel.active{display:block;}
.fg{margin-bottom:14px;}
label{display:block;font-size:12px;font-weight:600;color:#555;margin-bottom:5px;}
input{width:100%;height:44px;border:1.5px solid #e0e0e0;border-radius:10px;
      padding:0 14px;font-size:15px;outline:none;transition:border-color .2s;}
input:focus{border-color:#4CAF50;}
.btn{width:100%;height:46px;background:#2E7D32;color:white;border:none;
     border-radius:10px;font-size:16px;font-weight:700;cursor:pointer;margin-top:4px;}
.btn:hover{background:#1B5E20;}
.btn:disabled{background:#aaa;cursor:not-allowed;}
.msg{padding:12px 14px;border-radius:8px;font-size:14px;font-weight:600;
     margin-bottom:16px;display:none;text-align:center;}
.ok{background:#E8F5E9;color:#2E7D32;border:1px solid #A5D6A7;display:block;}
.er{background:#FFEBEE;color:#C62828;border:1px solid #EF9A9A;display:block;}
.footer{text-align:center;margin-top:18px;font-size:12px;color:#ccc;}
"""

APP_CSS = """
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
     background:#f5f5f0;min-height:100vh;}
.top{background:#2E7D32;color:white;padding:13px 20px;
     display:flex;align-items:center;justify-content:space-between;
     position:sticky;top:0;z-index:10;}
.tl{display:flex;align-items:center;gap:10px;}
.tl h1{font-size:17px;font-weight:700;}
.tl p{font-size:12px;opacity:.8;}
.tr{display:flex;gap:8px;}
.bxl{background:#FF8F00;color:white;border:none;padding:8px 14px;
     border-radius:8px;font-size:13px;font-weight:700;cursor:pointer;}
.bxl:hover{background:#E65100;}
.bout{background:rgba(255,255,255,.2);color:white;border:none;
      padding:8px 13px;border-radius:8px;font-size:13px;cursor:pointer;}
.bout:hover{background:rgba(255,255,255,.3);}
.wrap{max-width:780px;margin:0 auto;padding:20px 14px;}
.toast{position:fixed;top:14px;right:14px;padding:12px 16px;border-radius:9px;
       font-size:13px;font-weight:600;display:none;z-index:99;
       box-shadow:0 4px 14px rgba(0,0,0,.15);max-width:300px;}
.toast.show{display:flex;align-items:center;gap:7px;}
.tok{background:#E8F5E9;color:#1B5E20;border:1px solid #A5D6A7;}
.ter{background:#FFEBEE;color:#C62828;border:1px solid #EF9A9A;}
.stats{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:18px;}
.stat{background:white;border-radius:10px;padding:14px 16px;
      box-shadow:0 1px 3px rgba(0,0,0,.07);}
.sl{font-size:11px;color:#888;margin-bottom:4px;}
.sv{font-size:22px;font-weight:800;color:#2E7D32;}
.ss{font-size:11px;color:#bbb;}
.card{background:white;border-radius:12px;padding:18px 20px;
      box-shadow:0 1px 3px rgba(0,0,0,.07);margin-bottom:18px;}
.card h2{font-size:14px;font-weight:700;color:#333;margin-bottom:14px;
         display:flex;align-items:center;gap:8px;}
.card h2::before{content:'';width:4px;height:15px;background:#2E7D32;
                 border-radius:2px;display:inline-block;}
.gr3{display:grid;grid-template-columns:1fr 1fr auto;gap:10px;align-items:end;}
.gr4{display:grid;grid-template-columns:1fr 1fr 1fr auto;gap:8px;align-items:end;}
.fg{display:flex;flex-direction:column;gap:5px;}
label{font-size:12px;font-weight:600;color:#555;}
input[type=text],input[type=number],select{height:40px;border:1.5px solid #e0e0e0;
  border-radius:9px;padding:0 11px;font-size:14px;background:#fafafa;
  outline:none;transition:border-color .2s;width:100%;}
input:focus,select:focus{border-color:#4CAF50;background:white;}
.bs{height:40px;padding:0 16px;background:#2E7D32;color:white;border:none;
    border-radius:9px;font-size:14px;font-weight:700;cursor:pointer;white-space:nowrap;}
.bs:hover{background:#1B5E20;}
.bs:disabled{background:#aaa;cursor:not-allowed;}
.ba{height:40px;padding:0 14px;background:#1565C0;color:white;border:none;
    border-radius:9px;font-size:13px;font-weight:700;cursor:pointer;}
.ba:hover{background:#0D47A1;}
.bd{background:none;border:none;color:#EF5350;font-size:16px;cursor:pointer;
    padding:4px 8px;border-radius:6px;}
.bd:hover{background:#FFEBEE;}
.filters{display:flex;gap:8px;margin-bottom:12px;flex-wrap:wrap;}
.fb{padding:5px 13px;border:1.5px solid #ddd;border-radius:20px;background:white;
    font-size:12px;cursor:pointer;transition:all .15s;}
.fb:hover{border-color:#4CAF50;color:#2E7D32;}
.fb.active{background:#2E7D32;color:white;border-color:#2E7D32;}
table{width:100%;border-collapse:collapse;}
thead th{background:#f8f8f6;padding:10px 12px;text-align:left;font-size:11px;
         font-weight:700;color:#777;border-bottom:1px solid #eee;
         text-transform:uppercase;letter-spacing:.4px;}
tbody tr{border-bottom:.5px solid #f0f0ec;transition:background .1s;}
tbody tr:last-child{border-bottom:none;}
tbody tr:hover{background:#f9fdf9;}
tbody tr.hl{background:#E8F5E9!important;}
tbody td{padding:9px 12px;font-size:13px;}
.badge{display:inline-block;padding:2px 9px;border-radius:12px;
       font-size:11px;font-weight:700;}
.bv{background:#E3F2FD;color:#1565C0;}
.bm{background:#FFF8E1;color:#E65100;}
.bc{background:#F3E5F5;color:#6A1B9A;}
.fw{font-weight:700;}.mu{color:#aaa;font-size:11px;}
.empty{text-align:center;padding:30px;color:#bbb;font-size:14px;}
@media(max-width:600px){
  .gr3,.gr4{grid-template-columns:1fr;}
  .stats{grid-template-columns:1fr 1fr;}
}
"""

PAGE_LOGIN = f"""<!DOCTYPE html><html lang="fr"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>🐄 Gestion Ferme</title><style>{CSS}</style></head><body>
<div class="card">
  <div class="logo"><div class="em">🐄</div>
    <h1>Gestion Ferme</h1>
    <p class="sub">Suivi des poids des animaux</p>
  </div>
  <div class="tabs">
    <button class="tab active" id="t1" onclick="sw('login',this)">Se connecter</button>
    <button class="tab" id="t2" onclick="sw('register',this)">Créer un compte</button>
  </div>
  <div id="msg" class="msg"></div>

  <div id="login" class="panel active">
    <div class="fg"><label>Identifiant</label>
      <input id="lu" placeholder="ex: ahmed123" autocomplete="username"/></div>
    <div class="fg"><label>Mot de passe</label>
      <input id="lm" type="password" placeholder="••••••••"
        onkeydown="if(event.key==='Enter')login()"/></div>
    <button class="btn" id="bl" onclick="login()">Se connecter</button>
  </div>

  <div id="register" class="panel">
    <div class="fg"><label>Nom complet</label>
      <input id="rn" placeholder="ex: Ahmed Ben Ali"/></div>
    <div class="fg"><label>Identifiant (lettres/chiffres uniquement, sans espace)</label>
      <input id="ru" placeholder="ex: ahmed123"/></div>
    <div class="fg"><label>Mot de passe (minimum 4 caractères)</label>
      <input id="rm" type="password" placeholder="ex: 1234"/></div>
    <div class="fg"><label>Confirmer le mot de passe</label>
      <input id="rm2" type="password" placeholder="répète le mot de passe"
        onkeydown="if(event.key==='Enter')register()"/></div>
    <button class="btn" id="br" onclick="register()">Créer mon compte</button>
  </div>
  <div class="footer">Système de gestion des animaux • OEP</div>
</div>
<script>
function sw(id,btn){{
  ['login','register'].forEach(x=>document.getElementById(x).classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  btn.classList.add('active'); hide();
}}
function show(m,ok){{const e=document.getElementById('msg');e.textContent=m;e.className='msg '+(ok?'ok':'er');}}
function hide(){{document.getElementById('msg').className='msg';}}
function setBtn(id,loading,txt){{const b=document.getElementById(id);b.disabled=loading;b.textContent=loading?'⏳ Chargement...':txt;}}

async function login(){{
  const u=document.getElementById('lu').value.trim();
  const m=document.getElementById('lm').value;
  if(!u||!m){{show('⚠️ Remplis les deux champs.',false);return;}}
  setBtn('bl',true,'Se connecter'); hide();
  try{{
    const r=await post('/api/login',{{u,m}});
    if(r.ok){{
      localStorage.setItem('tk',r.tk);
      localStorage.setItem('nom',r.nom);
      localStorage.setItem('un',r.un);
      window.location.href='/app';
    }}else{{show('❌ '+r.msg,false);setBtn('bl',false,'Se connecter');}}
  }}catch(e){{show('❌ Erreur réseau — réessaie.',false);setBtn('bl',false,'Se connecter');}}
}}

async function register(){{
  const n=document.getElementById('rn').value.trim();
  const u=document.getElementById('ru').value.trim();
  const m=document.getElementById('rm').value;
  const m2=document.getElementById('rm2').value;
  if(!n||!u||!m||!m2){{show('⚠️ Remplis tous les champs.',false);return;}}
  if(m!==m2){{show('⚠️ Les mots de passe sont différents.',false);return;}}
  if(m.length<4){{show('⚠️ Mot de passe trop court (min. 4 caractères).',false);return;}}
  if(!/^[a-zA-Z0-9_]+$/.test(u)){{show('⚠️ Identifiant : lettres et chiffres seulement.',false);return;}}
  setBtn('br',true,'Créer mon compte'); hide();
  try{{
    const r=await post('/api/register',{{n,u,m}});
    if(r.ok){{
      sw('login',document.getElementById('t1'));
      document.getElementById('lu').value=u;
      show('✅ Compte créé ! Entre ton mot de passe et connecte-toi.',true);
    }}else show('❌ '+r.msg,false);
  }}catch(e){{show('❌ Erreur réseau — réessaie.',false);}}
  setBtn('br',false,'Créer mon compte');
}}

async function post(url,body){{
  const r=await fetch(url,{{method:'POST',
    headers:{{'Content-Type':'application/json'}},
    body:JSON.stringify(body)}});
  if(!r.ok)throw new Error('HTTP '+r.status);
  return r.json();
}}
</script></body></html>"""

PAGE_APP = f"""<!DOCTYPE html><html lang="fr"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>🐄 Mon Troupeau</title><style>{APP_CSS}</style></head><body>
<div class="top">
  <div class="tl"><span style="font-size:26px">🐄</span>
    <div><h1 id="tnm">Ferme</h1><p id="tsb">Mon troupeau</p></div></div>
  <div class="tr">
    <button class="bxl" onclick="excel()">📊 Excel</button>
    <button class="bout" onclick="logout()">⬅ Déconnexion</button>
  </div>
</div>
<div class="wrap">
  <div class="toast" id="toast"></div>
  <div class="stats" id="stats"></div>

  <div class="card">
    <h2>⚖️ Enregistrer une pesée</h2>
    <div class="gr3">
      <div class="fg"><label>Choisir l'animal</label>
        <select id="sel"><option value="">— Sélectionner —</option></select></div>
      <div class="fg"><label>Nouveau poids (kg)</label>
        <input type="number" id="inp" placeholder="ex: 465" min="1" max="2000" step="0.5"/></div>
      <div class="fg"><button class="bs" id="bpe" onclick="peser()">💾 Enregistrer</button></div>
    </div>
  </div>

  <div class="card">
    <h2>➕ Ajouter un animal</h2>
    <div class="gr4">
      <div class="fg"><label>Numéro (ex: ANI-001)</label>
        <input type="text" id="ai" placeholder="ANI-001"/></div>
      <div class="fg"><label>Espèce</label>
        <select id="ae"><option>Vache</option><option>Mouton</option><option>Chèvre</option></select></div>
      <div class="fg"><label>Poids initial (kg)</label>
        <input type="number" id="ap" placeholder="450" min="1"/></div>
      <div class="fg"><button class="ba" onclick="ajouter()">+ Ajouter</button></div>
    </div>
  </div>

  <div class="card">
    <h2>📋 Mon troupeau</h2>
    <div class="filters">
      <button class="fb active" onclick="filtre('tous',this)">Tous</button>
      <button class="fb" onclick="filtre('Vache',this)">🐄 Vaches</button>
      <button class="fb" onclick="filtre('Mouton',this)">🐑 Moutons</button>
      <button class="fb" onclick="filtre('Chèvre',this)">🐐 Chèvres</button>
    </div>
    <table><thead><tr>
      <th>N° Animal</th><th>Espèce</th><th>Poids</th><th>Dernière MAJ</th><th>Action</th>
    </tr></thead><tbody id="tb">
      <tr><td colspan="5" class="empty">⏳ Chargement...</td></tr>
    </tbody></table>
  </div>
</div>
<script>
const tk=localStorage.getItem('tk');
const nom=localStorage.getItem('nom');
const un=localStorage.getItem('un');
if(!tk){{window.location.href='/';}}
document.getElementById('tnm').textContent='Bonjour, '+nom+' 👋';

let data=[],flt='tous',hl=null;

async function api(method,url,body){{
  const o={{method,headers:{{'Content-Type':'application/json','X-Token':tk}}}};
  if(body)o.body=JSON.stringify(body);
  const r=await fetch(url,o);
  if(!r.ok)throw new Error('HTTP '+r.status);
  return r.json();
}}

async function charger(){{
  try{{
    const r=await api('GET','/api/animaux');
    data=r.animaux||[];
    selFill();stats();table(hl);
  }}catch(e){{
    document.getElementById('tb').innerHTML='<tr><td colspan="5" class="empty">❌ Erreur — actualise la page.</td></tr>';
  }}
}}

function selFill(){{
  const s=document.getElementById('sel');
  s.innerHTML='<option value="">— Sélectionner —</option>';
  data.forEach(a=>{{
    const o=document.createElement('option');
    o.value=a.id;o.textContent=a.id+' ('+a.espece+' — '+a.poids.toFixed(1)+' kg)';
    s.appendChild(o);
  }});
}}

function stats(){{
  const n=data.length,moy=n?(data.reduce((s,a)=>s+a.poids,0)/n).toFixed(1):0;
  const auj=new Date().toISOString().split('T')[0];
  const p=data.filter(a=>a.date===auj).length;
  document.getElementById('stats').innerHTML=`
    <div class="stat"><div class="sl">Mes animaux</div><div class="sv">${{n}}</div><div class="ss">total</div></div>
    <div class="stat"><div class="sl">Poids moyen</div><div class="sv">${{moy}}</div><div class="ss">kg</div></div>
    <div class="stat"><div class="sl">Pesées aujourd'hui</div><div class="sv">${{p}}</div><div class="ss">mises à jour</div></div>`;
}}

function table(h){{
  const f=flt==='tous'?data:data.filter(a=>a.espece===flt);
  const b=e=>e==='Vache'?'bv':e==='Mouton'?'bm':'bc';
  document.getElementById('tb').innerHTML=f.length?
    f.map(a=>`<tr class="${{a.id===h?'hl':''}}">
      <td class="fw">${{a.id}}</td>
      <td><span class="badge ${{b(a.espece)}}">${{a.espece}}</span></td>
      <td class="fw">${{a.poids.toFixed(1)}} kg</td>
      <td class="mu">${{a.date}}</td>
      <td><button class="bd" onclick="suppr('${{a.id}}')">🗑</button></td>
    </tr>`).join(''):
    '<tr><td colspan="5" class="empty">Aucun animal — ajoutes-en un avec le formulaire ci-dessus !</td></tr>';
}}

function filtre(f,btn){{
  flt=f;document.querySelectorAll('.fb').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');table(hl);
}}

function toast(msg,ok=true){{
  const t=document.getElementById('toast');
  t.textContent=msg;t.className='toast show '+(ok?'tok':'ter');
  setTimeout(()=>t.className='toast',4000);
}}

async function peser(){{
  const id=document.getElementById('sel').value;
  const p=parseFloat(document.getElementById('inp').value);
  if(!id){{toast('⚠️ Sélectionne un animal.',false);return;}}
  if(!p||p<=0){{toast('⚠️ Saisis un poids valide.',false);return;}}
  document.getElementById('bpe').disabled=true;
  try{{
    const r=await api('POST','/api/update',{{id,poids:p}});
    if(r.ok){{
      hl=id;await charger();
      const d=(p-r.ancien).toFixed(1);
      toast('✅ '+id+' : '+p.toFixed(1)+' kg ('+(d>=0?'+':'')+d+' kg)');
      document.getElementById('sel').value='';document.getElementById('inp').value='';
      setTimeout(()=>{{hl=null;table();}},4000);
    }}else toast('❌ Animal introuvable.',false);
  }}catch(e){{toast('❌ Erreur réseau.',false);}}
  document.getElementById('bpe').disabled=false;
}}

async function ajouter(){{
  const id=document.getElementById('ai').value.trim().toUpperCase();
  const esp=document.getElementById('ae').value;
  const p=parseFloat(document.getElementById('ap').value);
  if(!id){{toast('⚠️ Saisis un numéro (ex: ANI-001).',false);return;}}
  if(!p||p<=0){{toast('⚠️ Saisis un poids valide.',false);return;}}
  try{{
    const r=await api('POST','/api/ajouter',{{id,espece:esp,poids:p}});
    if(r.ok){{
      toast('✅ '+id+' ajouté avec succès !');
      document.getElementById('ai').value='';document.getElementById('ap').value='';
      await charger();
    }}else toast('❌ '+(r.msg||'Erreur'),false);
  }}catch(e){{toast('❌ Erreur réseau.',false);}}
}}

async function suppr(id){{
  if(!confirm('Supprimer '+id+' définitivement ?'))return;
  try{{
    const r=await api('POST','/api/supprimer',{{id}});
    if(r.ok){{toast('🗑 '+id+' supprimé.');await charger();}}
    else toast('❌ Erreur.',false);
  }}catch(e){{toast('❌ Erreur réseau.',false);}}
}}

async function excel(){{
  toast('⏳ Génération Excel...');
  try{{
    const r=await fetch('/api/excel',{{headers:{{'X-Token':tk}}}});
    if(!r.ok)throw new Error();
    const blob=await r.blob();
    const url=URL.createObjectURL(blob);
    const a=document.createElement('a');
    a.href=url;a.download='animaux_'+un+'_'+new Date().toISOString().split('T')[0]+'.xlsx';
    a.click();URL.revokeObjectURL(url);
    toast('📊 Excel téléchargé !');
  }}catch(e){{toast('❌ Erreur téléchargement.',false);}}
}}

function logout(){{if(confirm('Se déconnecter ?')){{localStorage.clear();window.location.href='/';}}}}
charger();
</script></body></html>"""

class H(BaseHTTPRequestHandler):
    def log_message(self, *a): pass
    def tk(self): return self.headers.get("X-Token","")
    def usr(self): return get_user(self.tk())
    def body(self):
        n=int(self.headers.get("Content-Length",0))
        return json.loads(self.rfile.read(n)) if n else {}
    def jout(self,d,s=200):
        b=json.dumps(d,ensure_ascii=False).encode()
        self.send_response(s);self.send_header("Content-Type","application/json;charset=utf-8")
        self.send_header("Content-Length",len(b));self.end_headers();self.wfile.write(b)
    def hout(self,h):
        b=h.encode();self.send_response(200)
        self.send_header("Content-Type","text/html;charset=utf-8")
        self.send_header("Content-Length",len(b));self.end_headers();self.wfile.write(b)

    def do_GET(self):
        p=urlparse(self.path).path; u=self.usr()
        if p=="/": self.hout(PAGE_LOGIN)
        elif p=="/app":
            if not u:
                self.send_response(302);self.send_header("Location","/");self.end_headers()
            else: self.hout(PAGE_APP)
        elif p=="/api/animaux":
            if not u: self.jout({"error":"non autorisé"},401);return
            d=load(DATA_FILE); self.jout({"animaux":d.get(u,[])})
        elif p=="/api/excel":
            if not u: self.jout({"error":"non autorisé"},401);return
            d=load(DATA_FILE); animaux=d.get(u,[])
            try:
                import openpyxl
                from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
                wb=openpyxl.Workbook();ws=wb.active;ws.title="Animaux"
                hf=PatternFill("solid",start_color="2E7D32")
                hfont=Font(bold=True,color="FFFFFF",name="Arial",size=11)
                thin=Border(left=Side(style="thin"),right=Side(style="thin"),
                            top=Side(style="thin"),bottom=Side(style="thin"))
                for c,h in enumerate(["N° Animal","Espèce","Poids (kg)","Dernière MAJ"],1):
                    cell=ws.cell(1,c,h);cell.font=hfont;cell.fill=hf
                    cell.alignment=Alignment(horizontal="center")
                for r,a in enumerate(animaux,2):
                    fill=PatternFill("solid",start_color="F1F8E9" if r%2==0 else "FFFFFF")
                    for c,v in enumerate([a["id"],a["espece"],a["poids"],a["date"]],1):
                        cell=ws.cell(r,c,v);cell.border=thin
                        cell.font=Font(name="Arial",size=10)
                        cell.fill=fill;cell.alignment=Alignment(horizontal="center")
                for col,w in zip("ABCD",[14,12,14,16]):ws.column_dimensions[col].width=w
                import tempfile
                tmp=tempfile.NamedTemporaryFile(suffix=".xlsx",delete=False)
                wb.save(tmp.name)
                with open(tmp.name,"rb") as f:content=f.read()
                os.unlink(tmp.name)
                self.send_response(200)
                self.send_header("Content-Type","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition",f'attachment; filename="animaux_{u}.xlsx"')
                self.send_header("Content-Length",len(content))
                self.end_headers();self.wfile.write(content)
            except Exception as e: self.jout({"error":str(e)},500)
        else: self.send_response(404);self.end_headers()

    def do_POST(self):
        p=urlparse(self.path).path; b=self.body()

        if p=="/api/register":
            n=b.get("n","").strip(); u=b.get("u","").strip(); m=b.get("m","")
            if not n or not u or not m:
                self.jout({"ok":False,"msg":"Remplis tous les champs."});return
            if not re.match(r'^[a-zA-Z0-9_]+$',u):
                self.jout({"ok":False,"msg":"Identifiant invalide."});return
            if len(m)<4:
                self.jout({"ok":False,"msg":"Mot de passe trop court."});return
            users=load(USERS_FILE)
            if u in users:
                self.jout({"ok":False,"msg":"Identifiant déjà utilisé."});return
            users[u]={"nom":n,"mdp":sha(m)}
            save(USERS_FILE,users)
            d=load(DATA_FILE);d[u]=[];save(DATA_FILE,d)
            self.jout({"ok":True})

        elif p=="/api/login":
            u=b.get("u","").strip(); m=b.get("m","")
            users=load(USERS_FILE); info=users.get(u,{})
            if info and info.get("mdp")==sha(m):
                t=new_session(u)
                self.jout({"ok":True,"tk":t,"nom":info["nom"],"un":u})
            else:
                self.jout({"ok":False,"msg":"Identifiant ou mot de passe incorrect."})

        elif p=="/api/update":
            u=self.usr()
            if not u: self.jout({"ok":False});return
            d=load(DATA_FILE)
            for a in d.get(u,[]):
                if a["id"]==b.get("id"):
                    ancien=a["poids"];a["poids"]=float(b["poids"])
                    a["date"]=date.today().isoformat()
                    save(DATA_FILE,d);self.jout({"ok":True,"ancien":ancien});return
            self.jout({"ok":False})

        elif p=="/api/ajouter":
            u=self.usr()
            if not u: self.jout({"ok":False});return
            d=load(DATA_FILE)
            if u not in d: d[u]=[]
            for a in d[u]:
                if a["id"]==b.get("id"):
                    self.jout({"ok":False,"msg":"Ce numéro existe déjà."});return
            d[u].append({"id":b.get("id","").upper(),"espece":b.get("espece","Vache"),
                         "poids":float(b.get("poids",0)),"date":date.today().isoformat()})
            save(DATA_FILE,d);self.jout({"ok":True})

        elif p=="/api/supprimer":
            u=self.usr()
            if not u: self.jout({"ok":False});return
            d=load(DATA_FILE); avant=len(d.get(u,[]))
            d[u]=[a for a in d.get(u,[]) if a["id"]!=b.get("id")]
            if len(d[u])<avant: save(DATA_FILE,d);self.jout({"ok":True})
            else: self.jout({"ok":False})
        else:
            self.send_response(404);self.end_headers()

if __name__=="__main__":
    init()
    PORT=int(os.environ.get("PORT",5000))
    server=HTTPServer(("",PORT),H)
    print(f"\n🌿 Application Ferme démarrée sur le port {PORT}")
    try: server.serve_forever()
    except KeyboardInterrupt: print("\n✅ Arrêtée.")
